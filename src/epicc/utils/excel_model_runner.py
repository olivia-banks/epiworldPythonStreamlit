"""
DEPRECATED, remove as per #26, "Remove XLSX Model Specification Code"

https://github.com/EpiForeSITE/epicc/issues/26
"""

from __future__ import annotations

import ast
import os
import re
from dataclasses import dataclass
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from epicc.utils.parameter_loader import flatten_dict

_CELL_REF_RE = re.compile(r"(\$?[A-Z]{1,3}\$?\d+)")
_RANGE_REF_RE = re.compile(r"(\$?[A-Z]{1,3}\$?\d+)\s*:\s*(\$?[A-Z]{1,3}\$?\d+)")


def _normalize_ref(ref: str) -> str:
    return ref.replace("$", "")


def _to_float(x: Any) -> float:
    if x is None or x == "":
        return 0.0
    if isinstance(x, (int, float)):
        return float(x)
    try:
        return float(str(x).strip())
    except Exception:
        return 0.0


def _round_if_number(value: Any) -> Any:
    """
    Tries to convert value to a float.
    - If magnitude > 10, round to nearest whole number (int).
    - If magnitude <= 10, round to 2 decimal places (float).
    Returns original value if conversion fails.
    """
    try:
        if isinstance(value, (int, float)):
            f_val = float(value)
        else:
            s_val = str(value).strip()
            if s_val == "":
                return value
            f_val = float(s_val)

        # --- CONDITIONAL ROUNDING LOGIC ---
        if abs(f_val) > 10:
            return int(round(f_val))
        else:
            return round(f_val, 2)

    except (ValueError, TypeError):
        return value


def _infer_indent_level(text: str) -> int:
    if not isinstance(text, str):
        return 0
    leading = len(text) - len(text.lstrip(" "))
    return leading // 2


def _excel_indent_level(cell) -> int:
    try:
        indent = getattr(cell.alignment, "indent", 0)
        if indent is None:
            indent = 0
        return int(indent)
    except Exception:
        return _infer_indent_level(str(cell.value) if cell.value else "")


def _col_to_index(col: str) -> int:
    col = col.upper().strip()
    n = 0
    for ch in col:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def _index_to_col(idx: int) -> str:
    result = ""
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        result = chr(rem + ord("A")) + result
    return result


def _scenario_columns_before_F(ws: Worksheet) -> list[str]:
    f_idx = _col_to_index("F")
    return [_index_to_col(i) for i in range(_col_to_index("B"), f_idx)]


_ALLOWED_AST_NODES = (
    ast.Expression,
    ast.BinOp,
    ast.UnaryOp,
    ast.Add,
    ast.Sub,
    ast.Mult,
    ast.Mod,
    ast.Div,
    ast.FloorDiv,
    ast.Pow,
    ast.USub,
    ast.UAdd,
    ast.Load,
    ast.Call,
    ast.Name,
    ast.Constant,
    ast.Compare,
    ast.Eq,
    ast.NotEq,
    ast.Gt,
    ast.GtE,
    ast.Lt,
    ast.LtE,
    ast.BoolOp,
    ast.And,
    ast.Or,
    ast.List,
    ast.Tuple,
)


def _safe_eval(expr: str, env: dict) -> Any:
    node = ast.parse(expr, mode="eval")
    for n in ast.walk(node):
        if not isinstance(n, _ALLOWED_AST_NODES):
            raise ValueError(f"Disallowed expression element: {type(n).__name__}")
        if isinstance(n, ast.Call):
            if not isinstance(n.func, ast.Name) or n.func.id not in env:
                raise ValueError("Only whitelisted functions are allowed in formulas")
        if isinstance(n, ast.Name):
            if n.id not in env:
                raise ValueError(f"Unknown name in expression: {n.id}")
    return eval(compile(node, "<excel_formula>", "eval"), {"__builtins__": {}}, env)


class ExcelValue:
    """Wrapper that handles Excel-style list broadcasting for operations"""

    def __init__(self, value):
        self.value = value

    def unwrap(self):
        return self.value

    def __add__(self, other):
        return self._binary_op(other, lambda a, b: a + b)

    def __radd__(self, other):
        return self._binary_op(other, lambda a, b: b + a)

    def __sub__(self, other):
        return self._binary_op(other, lambda a, b: a - b)

    def __rsub__(self, other):
        return self._binary_op(other, lambda a, b: b - a)

    def __mul__(self, other):
        return self._binary_op(other, lambda a, b: a * b)

    def __rmul__(self, other):
        return self._binary_op(other, lambda a, b: b * a)

    def __truediv__(self, other):
        return self._binary_op(other, lambda a, b: a / b if b != 0 else 0.0)

    def __rtruediv__(self, other):
        return self._binary_op(other, lambda a, b: b / a if a != 0 else 0.0)

    def __pow__(self, other):
        return self._binary_op(other, lambda a, b: a**b)

    def __rpow__(self, other):
        return self._binary_op(other, lambda a, b: b**a)

    def _binary_op(self, other, op):
        other_val = other.value if isinstance(other, ExcelValue) else other

        self_is_list = isinstance(self.value, list)
        other_is_list = isinstance(other_val, list)

        if self_is_list and other_is_list:
            return ExcelValue(
                [op(_to_float(a), _to_float(b)) for a, b in zip(self.value, other_val)]
            )
        elif self_is_list:
            return ExcelValue(
                [op(_to_float(a), _to_float(other_val)) for a in self.value]
            )
        elif other_is_list:
            return ExcelValue(
                [op(_to_float(self.value), _to_float(b)) for b in other_val]
            )
        else:
            return ExcelValue(op(_to_float(self.value), _to_float(other_val)))


@dataclass
class FormulaEngine:
    ws: Worksheet
    cache: dict[str, Any]

    def __init__(self, ws: Worksheet):
        self.ws = ws
        self.cache = {}

    def cell_value(self, ref: str) -> float:
        ref = _normalize_ref(ref)
        if ref in self.cache:
            return float(self.cache[ref])

        cell = self.ws[ref]
        v = cell.value

        if v is None or v == "":
            val = 0.0
        elif isinstance(v, (int, float)):
            val = float(v)
        elif isinstance(v, str) and v.startswith("="):
            try:
                out = self.eval_formula(v)
                if isinstance(out, list):
                    val = float(sum([_to_float(x) for x in out]))
                else:
                    val = float(out)
            except Exception as e:
                print(f"ERROR evaluating formula in {ref}: {v}")
                print(f"ERROR message: {e}")
                val = 0.0

        elif hasattr(v, "text") and isinstance(v.text, str):  # type: ignore
            # ArrayFormula support
            formula_text = v.text  # type: ignore
            if formula_text and formula_text.startswith("="):
                try:
                    out = self.eval_formula(formula_text)
                    if isinstance(out, list):
                        val = float(sum([_to_float(x) for x in out]))
                    else:
                        val = float(out)
                except Exception as e:
                    print(f"ERROR evaluating ArrayFormula in {ref}: {formula_text}")
                    print(f"ERROR message: {e}")
                    val = 0.0
            else:
                val = 0.0
        else:
            val = _to_float(v)

        self.cache[ref] = val
        return float(val)

    def eval_formula(self, formula: str) -> Any:
        f = formula.strip()
        if f.startswith("="):
            f = f[1:]

        # concat
        f = f.replace("&", "+")

        # ranges -> RANGE("A1","A10")
        def range_repl(m: re.Match) -> str:
            start = _normalize_ref(m.group(1))
            end = _normalize_ref(m.group(2))
            return f'RANGE("{start}","{end}")'

        expr = _RANGE_REF_RE.sub(range_repl, f)

        # cells -> VAL("A1") unless inside quotes
        def cell_repl(m: re.Match) -> str:
            ref = _normalize_ref(m.group(1))
            full_match_pos = m.start()
            quotes_before = expr[:full_match_pos].count('"')
            if quotes_before % 2 == 1:
                return m.group(0)
            return f'VAL("{ref}")'

        expr = _CELL_REF_RE.sub(cell_repl, expr)

        # comparisons
        expr = expr.replace("<>", "!=")
        expr = re.sub(r"(?<![<>=!])=(?![<>=])", "==", expr)

        expr = re.sub(r'(".*?")\s*\+\s*(VAL\(\"[A-Z]+\d+\"\))', r"\1 + STR(\2)", expr)

        expr = expr.replace("^", "**")

        # Wrap in EV
        expr = f"EV({expr})"

        # Excel functions
        def EV(x):
            if isinstance(x, ExcelValue):
                return x.unwrap()
            return x

        def IF(cond, a, b):
            return a if cond else b

        def STR(x):
            if isinstance(x, ExcelValue):
                x = x.unwrap()
            try:
                xf = float(_to_float(x))
                if abs(xf - round(xf)) < 1e-12:
                    return str(int(round(xf)))
                return str(xf)
            except Exception:
                return str(x)

        def _flatten_args(args):
            flat = []
            for a in args:
                if isinstance(a, ExcelValue):
                    a = a.unwrap()
                if isinstance(a, list):
                    flat.extend(a)
                else:
                    flat.append(a)
            return flat

        def SUM(*args):
            vals = _flatten_args(args)
            return float(sum([_to_float(v) for v in vals]))

        def MIN(*args):
            vals = _flatten_args(args)
            return float(min([_to_float(v) for v in vals])) if vals else 0.0

        def MAX(*args):
            vals = _flatten_args(args)
            return float(max([_to_float(v) for v in vals])) if vals else 0.0

        def RANGE(start_ref, end_ref):
            start_match = re.match(r"([A-Z]+)(\d+)", start_ref)
            end_match = re.match(r"([A-Z]+)(\d+)", end_ref)
            if not start_match or not end_match:
                return ExcelValue([])

            start_col = start_match.group(1)
            start_row = int(start_match.group(2))
            end_col = end_match.group(1)
            end_row = int(end_match.group(2))

            if start_col == end_col:
                values = []
                for row in range(start_row, end_row + 1):
                    values.append(self.cell_value(f"{start_col}{row}"))
                return ExcelValue(values)

            if start_row == end_row:
                values = []
                start_col_idx = _col_to_index(start_col)
                end_col_idx = _col_to_index(end_col)
                for col_idx in range(start_col_idx, end_col_idx + 1):
                    col = _index_to_col(col_idx)
                    values.append(self.cell_value(f"{col}{start_row}"))
                return ExcelValue(values)

            values = []
            start_col_idx = _col_to_index(start_col)
            end_col_idx = _col_to_index(end_col)
            for row in range(start_row, end_row + 1):
                for col_idx in range(start_col_idx, end_col_idx + 1):
                    col = _index_to_col(col_idx)
                    values.append(self.cell_value(f"{col}{row}"))
            return ExcelValue(values)

        def INDIRECT(x):
            if isinstance(x, ExcelValue):
                x = x.unwrap()
            if isinstance(x, (int, float)):
                return ExcelValue([])
            s = str(x).strip()
            m = re.search(r"(\d+)\s*:\s*([0-9]+(\.[0-9]+)?)", s)
            if not m:
                return ExcelValue([])
            start = int(float(m.group(1)))
            end = int(float(m.group(2)))
            if end < start:
                start, end = end, start
            return ExcelValue(list(range(start, end + 1)))

        def ROW(v):
            if isinstance(v, ExcelValue):
                v = v.unwrap()
            if isinstance(v, list):
                return ExcelValue([int(x) for x in v])
            if isinstance(v, (int, float)):
                return ExcelValue([int(v)])
            try:
                return ExcelValue([int(float(str(v)))])
            except Exception:
                return ExcelValue([])

        def SUMPRODUCT(*args):
            if not args:
                return 0.0

            unwrapped = []
            for a in args:
                if isinstance(a, ExcelValue):
                    unwrapped.append(a.unwrap())
                else:
                    unwrapped.append(a)

            if len(unwrapped) == 1 and isinstance(unwrapped[0], list):
                return float(sum([_to_float(v) for v in unwrapped[0]]))

            lists: list[list[float]] = []
            for a in unwrapped:
                if isinstance(a, list):
                    lists.append([_to_float(v) for v in a])
                else:
                    lists.append([_to_float(a)])

            max_len = max(len(L) for L in lists)

            norm: list[list[float]] = []
            for L in lists:
                if len(L) == 1 and max_len > 1:
                    norm.append(L * max_len)
                else:
                    norm.append(L)

            total = 0.0
            for i in range(max_len):
                prod = 1.0
                for L in norm:
                    prod *= float(_to_float(L[i]))
                total += prod

            return float(total)

        def VAL(r):
            return ExcelValue(self.cell_value(r))

        env = {
            "VAL": VAL,
            "EV": EV,
            "IF": IF,
            "SUM": SUM,
            "MIN": MIN,
            "MAX": MAX,
            "RANGE": RANGE,
            "INDIRECT": INDIRECT,
            "ROW": ROW,
            "SUMPRODUCT": SUMPRODUCT,
            "STR": STR,
            "ExcelValue": ExcelValue,
        }

        out = _safe_eval(expr, env)

        if isinstance(out, ExcelValue):
            out = out.unwrap()

        if isinstance(out, list):
            return [float(_to_float(v)) for v in out]
        return float(_to_float(out))


def excel_rows_to_nested_dict(rows: list[tuple[int, str, Any]]) -> dict:
    root: dict[str, Any] = {}
    stack: list[tuple[int, dict[str, Any]]] = [(0, root)]

    for level, name, value in rows:
        while len(stack) > 1 and stack[-1][0] >= level + 1:
            stack.pop()
        parent = stack[-1][1]
        if value is None:
            parent[name] = {}
            stack.append((level + 1, parent[name]))
        else:
            parent[name] = value
    return root


def apply_params_to_workbook(
    ws: Worksheet,
    params: dict[str, Any],
    name_col: str = "F",
    value_col: str = "G",
    start_row: int = 3,
    overwrite_formulas: bool = True,
):
    lookup: dict[str, Any] = {}
    for k, v in params.items():
        norm = str(k).replace("\t", "").strip()
        lookup[norm] = v

    for r in range(start_row, ws.max_row + 1):
        name_cell = ws[f"{name_col}{r}"]
        val_cell = ws[f"{value_col}{r}"]

        if name_cell.value is None:
            continue

        name = str(name_cell.value).strip()
        if not name:
            continue

        if (
            isinstance(val_cell.value, str) and str(val_cell.value).startswith("=")
        ) and not overwrite_formulas:
            continue

        if (
            name in lookup
            and lookup[name] is not None
            and str(lookup[name]).strip() != ""
        ):
            ws[f"{value_col}{r}"].value = lookup[name]


def load_excel_params_defaults_with_computed(
    excel_file,
    sheet_name: str | None = None,
    name_col: str = "F",
    value_col: str = "G",
    start_row: int = 3,
) -> tuple[dict[str, Any], dict[str, Any]]:
    wb = load_workbook(excel_file, data_only=False)
    ws = wb[sheet_name] if sheet_name else wb.active
    assert ws

    engine = FormulaEngine(ws)

    editable_rows: list[tuple[int, str, Any]] = []
    computed_defaults: dict[str, Any] = {}

    for r in range(start_row, ws.max_row + 1):
        name_cell = ws[f"{name_col}{r}"]
        val_cell = ws[f"{value_col}{r}"]

        if name_cell.value is None:
            continue

        raw_name = str(name_cell.value).strip()
        if raw_name == "":
            continue

        level = _excel_indent_level(name_cell)
        raw_val = val_cell.value

        if raw_val is None or str(raw_val).strip() == "":
            editable_rows.append((level, raw_name, None))
            continue

        if isinstance(raw_val, str) and raw_val.startswith("="):
            evaluated = engine.cell_value(f"{value_col}{r}")
            editable_rows.append((level, raw_name, evaluated))
            continue

        editable_rows.append((level, raw_name, raw_val))

    nested = excel_rows_to_nested_dict(editable_rows)
    editable_defaults = flatten_dict(nested)

    return editable_defaults, computed_defaults


def _find_outcome_header_row(ws: Worksheet) -> int | None:
    """
    Looks for the start of the result table by finding the first
    non-empty cell in Column A, starting from Row 2.
    """
    for r in range(2, ws.max_row + 1):
        a = ws[f"A{r}"].value
        if a is not None and str(a).strip() != "":
            return r
    return None


def _iter_outcome_rows(
    ws: Worksheet, header_row: int, scenario_cols: list[str]
) -> list[int]:
    rows: list[int] = []
    blank_streak = 0
    r = header_row + 1

    while r <= ws.max_row:
        a_val = ws[f"A{r}"].value
        a_text = "" if a_val is None else str(a_val).strip()

        has_any = False
        if a_text != "":
            has_any = True
        else:
            for col in scenario_cols:
                v = ws[f"{col}{r}"].value
                if v is not None and str(v).strip() != "":
                    has_any = True
                    break

        if not has_any:
            blank_streak += 1
            if blank_streak >= 3:
                break
        else:
            blank_streak = 0
            rows.append(r)

        r += 1
    return rows


def _detect_active_scenario_columns(
    ws: Worksheet, header_row: int, scenario_cols: list[str], rows: list[int]
) -> list[str]:
    active: list[str] = []
    for col in scenario_cols:
        header = ws[f"{col}{header_row}"].value
        has_header = header is not None and str(header).strip() != ""
        has_data = False
        for r in rows:
            v = ws[f"{col}{r}"].value
            if v is not None and str(v).strip() != "":
                has_data = True
                break
        if has_header or has_data:
            active.append(col)
    return active


def get_scenario_headers(excel_file, sheet_name: str | None = None) -> dict[str, str]:
    """
    Reads the header row (found by _find_outcome_header_row) and extracts
    current values for columns B, C, D, E.
    """
    wb = load_workbook(excel_file, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    assert ws

    header_row = _find_outcome_header_row(ws)

    if header_row is None:
        header_row = 1

    cols = _scenario_columns_before_F(ws)

    headers = {}
    for col in cols:
        val = ws[f"{col}{header_row}"].value
        if val is not None and str(val).strip() != "":
            headers[col] = str(val).strip()

    return headers


def build_sections_from_excel_outcomes(
    ws: Worksheet,
    engine: FormulaEngine,
    header_row: int,
    label_overrides: dict[str, str] | None = None,
) -> list[dict]:
    scenario_cols_all = _scenario_columns_before_F(ws)
    outcome_rows = _iter_outcome_rows(ws, header_row, scenario_cols_all)
    scenario_cols = _detect_active_scenario_columns(
        ws, header_row, scenario_cols_all, outcome_rows
    )

    if label_overrides is None:
        label_overrides = {}

    header_cell_val = ws[f"A{header_row}"].value
    first_col_title = str(header_cell_val).strip() if header_cell_val else "Outcome"

    col_titles = {}
    for col in scenario_cols:
        if col in label_overrides and label_overrides[col].strip() != "":
            col_titles[col] = label_overrides[col].strip()
        else:
            val = ws[f"{col}{header_row}"].value
            col_titles[col] = (
                str(val).strip() if val is not None and str(val).strip() != "" else col
            )

    sections: list[dict] = []
    current_title: str | None = None
    current_records: list[dict] = []

    for r in outcome_rows:
        a_val = ws[f"A{r}"].value
        a_text = "" if a_val is None else str(a_val).strip()
        if a_text == "":
            continue

        all_blank = True
        for col in scenario_cols:
            v = ws[f"{col}{r}"].value
            if v is not None and str(v).strip() != "":
                all_blank = False
                break

        if all_blank:
            if current_title and current_records:
                sections.append(
                    {"title": current_title, "content": [pd.DataFrame(current_records)]}
                )
                current_records = []
            current_title = a_text
            continue

        rec = {first_col_title: a_text}
        for col in scenario_cols:
            raw_val = engine.cell_value(f"{col}{r}")
            rec[col_titles[col]] = _round_if_number(raw_val)
        current_records.append(rec)

    if current_title and current_records:
        sections.append(
            {"title": current_title, "content": [pd.DataFrame(current_records)]}
        )

    if not sections and current_records:
        sections = [{"title": "Results", "content": [pd.DataFrame(current_records)]}]

    return sections


def _is_numberish(v: Any) -> bool:
    if v is None or v == "":
        return False
    if isinstance(v, (int, float)):
        return True
    if isinstance(v, str) and v.strip().startswith("="):
        return True
    try:
        float(str(v).strip())
        return True
    except Exception:
        return False


def _cell_display(ws: Worksheet, engine: FormulaEngine, cell_ref: str) -> Any:
    v = ws[cell_ref].value
    if _is_numberish(v):
        val = engine.cell_value(cell_ref)
        return _round_if_number(val)
    if v is None:
        return ""
    return str(v).strip()


def build_sections_from_generic_tables(
    ws: Worksheet, engine: FormulaEngine
) -> list[dict]:
    max_scan_rows = min(ws.max_row, 250)
    max_scan_cols = _col_to_index("E")

    def cell_is_blank(v: Any) -> bool:
        return v is None or str(v).strip() == ""

    tables: list[tuple[int, int, int, int]] = []

    for top in range(1, max_scan_rows + 1):
        for left_idx in range(1, max_scan_cols - 1):
            label = ws[f"{_index_to_col(left_idx)}{top}"].value
            if cell_is_blank(label):
                continue

            numeric_cols = 0
            for j in range(left_idx + 1, max_scan_cols + 1):
                v = ws[f"{_index_to_col(j)}{top}"].value
                if _is_numberish(v):
                    numeric_cols += 1
            if numeric_cols < 2:
                continue

            bottom = top
            while bottom + 1 <= max_scan_rows:
                next_label = ws[f"{_index_to_col(left_idx)}{bottom + 1}"].value
                if cell_is_blank(next_label):
                    break
                has_num = False
                for j in range(left_idx + 1, max_scan_cols + 1):
                    v = ws[f"{_index_to_col(j)}{bottom + 1}"].value
                    if _is_numberish(v):
                        has_num = True
                        break
                if not has_num:
                    break
                bottom += 1

            right = left_idx
            for j in range(left_idx + 1, max_scan_cols + 1):
                col_has_any = False
                for rr in range(top, bottom + 1):
                    v = ws[f"{_index_to_col(j)}{rr}"].value
                    if not cell_is_blank(v):
                        col_has_any = True
                        break
                if col_has_any:
                    right = j

            if bottom - top + 1 >= 2 and right - left_idx >= 1:
                tables.append((top, left_idx, bottom, right))

    if not tables:
        return [
            {
                "title": "Outputs",
                "content": [
                    pd.DataFrame(
                        [
                            {
                                "Error": "No Outcome found and no output table detected in A–E."
                            }
                        ]
                    )
                ],
            }
        ]

    tables.sort(key=lambda t: (t[2] - t[0] + 1) * (t[3] - t[1] + 1), reverse=True)
    top, left_idx, bottom, right_idx = tables[0]
    header_row = top

    headers = []
    for c in range(left_idx, right_idx + 1):
        v = ws[f"{_index_to_col(c)}{header_row}"].value
        headers.append(
            str(v).strip()
            if v is not None and str(v).strip() != ""
            else _index_to_col(c)
        )

    records = []
    for r in range(header_row + 1, bottom + 1):
        row_obj = {}
        for c, h in zip(range(left_idx, right_idx + 1), headers):
            cell_ref = f"{_index_to_col(c)}{r}"
            row_obj[h] = _cell_display(ws, engine, cell_ref)
        records.append(row_obj)

    df = pd.DataFrame(records)

    def col_is_effectively_empty(series: pd.Series) -> bool:
        for x in series.tolist():
            if x is None:
                continue
            s = str(x).strip()
            if s == "":
                continue
            if not _is_numberish(x):
                return False
            if abs(_to_float(x)) > 1e-12:
                return False
        return True

    df = df.loc[:, [c for c in df.columns if not col_is_effectively_empty(df[c])]]

    return [{"title": "Outputs", "content": [df]}]


# Main


def run_excel_driven_model(
    excel_file,
    filename: str,
    params: dict[str, Any],
    sheet_name: str | None = None,
    label_overrides: dict[str, str] | None = None,
) -> dict[str, Any]:
    wb = load_workbook(excel_file, data_only=False)
    ws = wb[sheet_name] if sheet_name else wb.active
    assert ws

    apply_params_to_workbook(ws, params, start_row=3, overwrite_formulas=True)

    engine = FormulaEngine(ws)
    header_row = _find_outcome_header_row(ws)

    if header_row is not None:
        sections = build_sections_from_excel_outcomes(
            ws, engine, header_row, label_overrides
        )
    else:
        sections = build_sections_from_generic_tables(ws, engine)

    model_name = os.path.splitext(os.path.basename(filename))[0]

    return {
        "model_title": model_name,
        "model_description": "Excel-driven model",
        "sections": sections,
    }
