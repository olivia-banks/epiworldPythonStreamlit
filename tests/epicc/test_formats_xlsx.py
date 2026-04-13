from io import BytesIO

import openpyxl
import pytest
from pydantic import BaseModel, Field

from epicc.formats.xlsx import XLSXFormat, _field_descriptions


def _make_xlsx(rows: list[list]) -> BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws

    for row in rows:
        ws.append(row)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return buf


def _fmt() -> XLSXFormat:
    return XLSXFormat("test.xlsx")


def test_read_dot_notation_creates_nested():
    buf = _make_xlsx(
        [
            ["param", "value"],
            ["costs.latent", 300],
            ["costs.active", 500],
        ]
    )
    data, _ = _fmt().read(buf)
    assert data == {"costs": {"latent": 300, "active": 500}}


def test_read_skips_empty_rows():
    buf = _make_xlsx([["param", "value"], ["a", 1], [None, None], ["b", 2]])
    data, _ = _fmt().read(buf)
    assert data == {"a": 1, "b": 2}


def test_read_too_few_columns_raises():
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws

    ws.append(["only_one_col"])
    ws.append(["value"])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    with pytest.raises(ValueError, match="at least 2 columns"):
        _fmt().read(buf)


def test_write_with_template_round_trips():
    buf = _make_xlsx([["param", "value"], ["x", 1], ["y", 2]])
    _, wb_template = _fmt().read(buf)

    result_bytes = _fmt().write({"x": 99}, wb_template)

    wb_out = openpyxl.load_workbook(BytesIO(result_bytes))
    ws_out = wb_out.active
    assert ws_out
    rows = list(ws_out.iter_rows(values_only=True))
    row_dict = {r[0]: r[1] for r in rows[1:] if r[0] is not None}
    assert row_dict["x"] == 99


def test_write_with_template_supports_nested_data():
    buf = _make_xlsx(
        [
            ["param", "value"],
            ["costs.latent", 300],
            ["costs.active", 500],
        ]
    )
    _, wb_template = _fmt().read(buf)

    result_bytes = _fmt().write({"costs": {"latent": 1}}, wb_template)

    wb_out = openpyxl.load_workbook(BytesIO(result_bytes))
    ws_out = wb_out.active
    assert ws_out
    rows = list(ws_out.iter_rows(values_only=True))
    row_dict = {r[0]: r[1] for r in rows[1:] if r[0] is not None}
    assert row_dict["costs.latent"] == 1
    assert row_dict["costs.active"] == 500


def test_write_without_template_not_empty():
    """Regression: write() with no template previously produced an empty workbook."""
    result_bytes = _fmt().write({"a": 1, "b": 2})

    assert len(result_bytes) > 0
    wb_out = openpyxl.load_workbook(BytesIO(result_bytes))
    ws_out = wb_out.active
    assert ws_out
    rows = list(ws_out.iter_rows(values_only=True))
    # Header row + at least one data row
    assert len(rows) >= 2
    row_dict = {r[0]: r[1] for r in rows[1:] if r[0] is not None}
    assert row_dict == {"a": 1, "b": 2}


def test_write_without_template_supports_nested_data():
    """Regression: nested dicts should be flattened to dot-notation when no template provided."""
    result_bytes = _fmt().write({"costs": {"latent": 300, "active": 500}})

    wb_out = openpyxl.load_workbook(BytesIO(result_bytes))
    ws_out = wb_out.active
    assert ws_out
    rows = list(ws_out.iter_rows(values_only=True))
    row_dict = {r[0]: r[1] for r in rows[1:] if r[0] is not None}
    assert row_dict == {"costs.latent": 300, "costs.active": 500}


class _SimpleModel(BaseModel):
    alpha: int = Field(1, description="Alpha value")
    beta: str = Field("x", description="Beta value")


def test_field_descriptions_flat():
    result = _field_descriptions(_SimpleModel)
    assert result == {"alpha": "Alpha value", "beta": "Beta value"}


def test_write_without_template_emits_descriptions():
    """Column C should contain field descriptions when pydantic_model is provided."""
    result_bytes = _fmt().write(
        {"alpha": 1, "beta": "x"},
        pydantic_model=_SimpleModel,
    )

    wb_out = openpyxl.load_workbook(BytesIO(result_bytes))
    ws_out = wb_out.active
    assert ws_out
    rows = list(ws_out.iter_rows(values_only=True))
    # Header row
    assert rows[0] == ("Parameter", "Value", "Description")
    desc_dict = {r[0]: r[2] for r in rows[1:] if r[0] is not None}
    assert desc_dict["alpha"] == "Alpha value"
    assert desc_dict["beta"] == "Beta value"


def test_write_without_template_no_pydantic_model_description_empty():
    """Without pydantic_model, description column should be empty strings."""
    result_bytes = _fmt().write({"x": 1})

    wb_out = openpyxl.load_workbook(BytesIO(result_bytes))
    ws_out = wb_out.active
    assert ws_out
    rows = list(ws_out.iter_rows(values_only=True))
    assert rows[0] == ("Parameter", "Value", "Description")
    assert rows[1][2] in (None, "")
